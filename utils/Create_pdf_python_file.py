# %%
# We can't import any self made functions because this uses a different intrepreter

# %%
from fpdf import FPDF

pdf = FPDF()


# %%
def text_to_pdf():
 
    # Stats section
    from fpdf import FPDF

    def PDF_Char_Sheet():
        pdf = FPDF()
        pdf.add_page()
        
        #Character Name 
        pdf.set_font("Arial", size=11)
        pdf.text(10, 10, "Name: (Name_String)")   




        # Ability scores title
        pdf.set_font("Arial", size=20)
        pdf.text(20, 20, "Ability Scores:")

        # Add text inside the ellipse
        pdf.set_font("Arial", size=12)
    
        
        # Drawing all rectangles
        # Ability Score Rectangles
        
        pdf.rect(25, 25, 15, 10,style='F')
        pdf.set_fill_color(0, 0, 0)
        pdf.rect(40, 25, 15, 10,style='F')
        pdf.set_fill_color(0, 0, 0)
        pdf.rect(25, 35, 15, 10,style='F')
        pdf.set_fill_color(0, 0, 0)
        pdf.rect(40, 35, 15, 10,style='F')
        pdf.set_fill_color(0, 0, 0)
        pdf.rect(25, 45, 15, 10,style='F')
        pdf.set_fill_color(0, 0, 0)
        pdf.rect(40, 45, 15, 10,style='F')
        pdf.set_fill_color(0, 0, 0)
        pdf.rect(25, 55, 15, 10,style='F')
        pdf.set_fill_color(0, 0, 0)
        pdf.rect(40, 55, 15, 10,style='F')
        pdf.set_fill_color(0, 0, 0)
        pdf.rect(25, 65, 15, 10,style='F')
        pdf.set_fill_color(0, 0, 0)
        pdf.rect(40, 65, 15, 10,style='F')
        pdf.set_fill_color(0, 0, 0)
        pdf.rect(25, 75, 15, 10,style='F')
        pdf.set_fill_color(0, 0, 0)
        pdf.rect(40, 75, 15, 10,style='F')
        pdf.set_fill_color(0, 0, 0)

        #Personality Trait Rectangle
        pdf.set_line_width(1)
        pdf.rect(75, 20, 80, 40) # This creates the box around it
        pdf.set_fill_color(0, 255, 0)  # Green fill color
        pdf.rect(75, 20, 80, 40, style = 'F') # This adds in the filler color
        pdf.set_fill_color(255, 255, 0)  # Yellow fill color
        
        # Add text inside the rectangle
        pdf.set_text_color(255, 255, 255)  # Text color (black in this example)
        pdf.text(30, 25, "STR")
        pdf.text(45, 25, "20")
        pdf.text(30, 35, "DEX:")
        pdf.text(45, 35, "10")

        pdf.text(30, 45, "CON:")
        pdf.text(45, 45, "10")
        pdf.text(30, 55, "WIS:")
        pdf.text(45, 55, "10")

        pdf.text(30, 65, "INT:")
        pdf.text(45, 65, "10")
        pdf.text(30, 75, "CHA:")
        pdf.text(45, 75, "10")

        # Name
        pdf.set_font("Arial", size=20)
        pdf.text(80, 15, "Name: (Name_string_here) ")
        # Personality Traits
        pdf.text(80, 30, "Personality Traits: ")     
        pdf.set_font("Arial", size=10)
        pdf.text(80, 35, "Personality Trait 1 ")   
        pdf.text(80, 40, "Personality Trait 2 ")
        pdf.text(80, 45, "Personality Trait 3 ")
        pdf.text(80, 50, "Personality Trait 4 ")
        pdf.text(80, 55, "Personality Trait 5 ")           
        # Ability Traits

        # Class (+ Class Levels + Archetype + HP)
        # Appearance
        # Profession
        # Mannerisms
        # Alignment + Deity
        # Languages
        # Skills (Primary Class)
        # Skills (Secondary Class)



        pdf.output('PDF_Char_Sheet.pdf')

    if __name__ == '__main__':
        PDF_Char_Sheet()

text_to_pdf()






# %%
text_to_pdf()

# %%




#incorrectly reads the PDF and inputs in a different area (probably just a better idea to create a custom PDF with the data filled in)

# %%

# %%


import openpyxl
import re

def text_to_excel(text_file, excel_file):
  """Converts a text file to an Excel file.

  Args:
    text_file: The path to the text file.
    excel_file: The path to the Excel file.
  """

  # Create an Excel workbook.
  wb = openpyxl.Workbook()

  # Create a worksheet.
  ws = wb.active

  # Open the text file.
  with open(text_file, "r") as f:
    # Read the text file line by line.
    for line in f:
      # Split the line into columns.
      columns = re.split(r'\s+', line)

      # Write the columns to the Excel worksheet.
      for column in columns:
        ws.cell(row=ws.max_row + 1, column=column).value = column

  # Save the Excel workbook.
  wb.save(excel_file)

# Example usage:
text_file = "/path/to/text/file.txt"
excel_file = "/path/to/excel/file.xlsx"

text_to_excel(text_file, excel_file)
# %%
